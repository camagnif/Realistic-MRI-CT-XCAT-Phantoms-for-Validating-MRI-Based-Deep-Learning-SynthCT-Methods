# -*- coding: utf-8 -*-
"""
Created on Sat Jun  8 15:56:14 2024

@author: fanciicam
"""

'''METRICS COMPUTATION'''

#______________________________LIBRARIES_and_FOLDERS___________________________

import os
import SimpleITK as sitk
import numpy as np
import tensorflow as tf
from image_similarity_measures import quality_metrics
import cv2
import matplotlib.pyplot as plt
from tensorflow.keras import layers
import tensorflow_addons as tfa
from tensorflow import keras
import openpyxl
from openpyxl.styles import Alignment
import glob
from scipy.fft import fft2
from scipy.ndimage import gaussian_filter
from scipy.stats import pearsonr
from random import randint


#_____________________________________FOR METRICS______________________________
# PAIRED
desired = 'test'

if desired=='train':
    #id_ph = [50, 76, 77, 80, 89, 92, 96, 98, 106, 108, 117, 118, 128, 139, 140, 141, 142, 144, 145, 146, 147, 148, 149, 150, 151, 152, 153, 154, 155, 157, 159, 162, 163, 164, 166, 167, 168, 169, 170, 171, 173, 175, 176, 178, 180, 182, 184, 196, 200, 201]
    #id_ph = [77, 89, 92, 96, 106, 118, 139, 140, 141, 142, 145, 146, 149, 150, 151, 152, 153, 154, 155, 157, 162, 166, 168, 169, 170, 171, 175, 178, 182, 196] # 30_cropped
    id_ph = [108, 117, 128, 144, 147, 148, 159, 163, 164, 167, 173, 176, 180, 184, 200, 201, 76, 80, 98] # additional 19 of training resized to 320x260
    #folder_COMBAT = '//NAS-CARTCAS/camagnif/tesi/Data/mri/vibe_combat/phantoms_new/phantoms_cut/VIBE/VIBE_noise/VIBE_recon/VIBE_denoised_clean_matched/'
    folder_COMBAT = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/VALIDATION/train_originalph_320x260_30_NOUSED/new/'
    #folder_MRI_corrected = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/MRI/train/train_348x348_50/'
    folder_MRI_corrected = '//NAS-CARTCAS/camagnif/tesi/Data/mri/results_revisori/7th/test/inference/train_volumes/'
    #folder_Mask = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/MRI/train/train_348x348_mask/'
    folder_Mask = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/VALIDATION/320x260_mask_35/new/'
elif desired=='test':
    # id_ph = [50, 71, 86, 99, 143]
    # folder_COMBAT = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/MRI/test/test_originalph_348x348_5/'
    # folder_MRI_corrected = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/MRI/test/test_348x348_5/'
    # folder_Mask = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/MRI/test/test_348x348_mask/'
    id_ph = [50, 51, 71, 86, 99, 143]
    folder_COMBAT = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/VALIDATION/train_originalph_320x260_30_NOUSED/new/test/'
    folder_MRI_corrected = '//NAS-CARTCAS/camagnif/tesi/Data/mri/results_revisori/5th/test/inference/test_volumes/'
    folder_Mask = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/VALIDATION/320x260_mask_35/new/test/'

# UNPAIRED
folder_unpaired = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/MRI/test/Unpaired/'

#_____________________________________HISTOGRAM________________________________

folder_all_fake = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/MRI/test/Unpaired/MRI_fake_all/'
folder_all_pz = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/MRI/test/Unpaired/MRI_real_all/'
folder_phantom = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/MRI/test/Unpaired/Phantoms_recon_all/'
folder = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/MRI/test/'

#_________________________________METRICS ON PATCHES___________________________

folder_weights = '//NAS-CARTCAS/camagnif/tesi/Data/mri/weights_TRAINING_MRI_DEF/'
folder_results_patches = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/MRI/test/'

#%%

#____________________________________FUNCTIONS_________________________________

def load(fileName):
    reader = sitk.ImageFileReader() # reads an image file and returns an SItkimage
    reader.SetImageIO("MetaImageIO")
    reader.SetFileName(fileName)
    imageSitk = reader.Execute()
    imageNp = sitk.GetArrayViewFromImage(imageSitk) 
    imageNp = np.moveaxis(imageNp, 0, 2) 
    tensorImg = tf.convert_to_tensor(imageNp, dtype=tf.float32)

    return tensorImg

# SSIM
def ssim(image_pred, image_gt, range):
    return quality_metrics.ssim(image_pred, image_gt, range)

# FSIM
def fsim(image_pred, image_gt, range):
    fsim = quality_metrics.fsim(image_pred, image_gt, range)
    return fsim

# EPR and EGR
def edge_ratios(img1, img2, threshold=0.8):

    # edge extraction
    sobel_filter = tf.constant([[1, 2, 1], [0, 0, 0], [-1, -2, -1]], dtype=tf.float32)
    img1_edges = tf.nn.conv2d(np.expand_dims(img1, 0), tf.expand_dims(tf.expand_dims(sobel_filter, -1), -1), strides=[1, 1, 1, 1], padding='SAME')
    img2_edges = tf.nn.conv2d(np.expand_dims(img2, 0), tf.expand_dims(tf.expand_dims(sobel_filter, -1), -1), strides=[1, 1, 1, 1], padding='SAME')

    # edges are positive and negative
    edge_gt = np.abs(tf.squeeze(img1_edges))
    edge_pred = np.abs(tf.squeeze(img2_edges))
    _, binary_img_gt = cv2.threshold(edge_gt, threshold, 1, cv2.THRESH_BINARY)
    _, binary_img_pred = cv2.threshold(edge_pred, threshold, 1, cv2.THRESH_BINARY)
    inter = cv2.bitwise_and(binary_img_gt, binary_img_pred) # edge map in common between edges1 and edges2
    inter = np.abs(inter)
    _, binary_img_inter = cv2.threshold(inter, threshold * 0.625, 1, cv2.THRESH_BINARY)
    
    # compute number of pixels = 1 belonging to edges
    tot_gt = np.sum(binary_img_gt)
    tot_pred = np.sum(binary_img_pred)
    tot_inter = np.sum(binary_img_inter)

    # EPR and EGR
    epr = tot_inter / tot_gt
    egr = tot_pred / tot_gt

    return epr, egr

# extract from volume the slice i 
def extract_slice(volume, i):
    slice_out = tf.slice(volume, [0, 0, i], [volume.shape[0], volume.shape[1], 1])
    return slice_out

def normalize(input_image):
    input_image = 2*((input_image-np.min(input_image)) / (np.max(input_image)-np.min(input_image))) - 1
    return input_image

# mean absolute error
def mean_absolute_error(image_pred, image_gt, mask):
    mae = np.mean(np.abs(image_pred[mask==1] - image_gt[mask==1]))
    return mae

def save_xlsx_paired(path, id_phantoms, MAE_list, SSIM_list, FSIM_list, EPR_list, EGR_list):
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet = wb_obj.active  # Access the active sheet
    max_column = sheet.max_column
    max_row = sheet.max_row
    dim = len(id_phantoms)

    for j in range(1, max_column + 1):
        HEADER = sheet.cell(row=1, column=j)

        if HEADER.value == "ID":
            for i in range(dim):
                cell = sheet.cell(max_row + 1 + i, j)
                cell.value = id_phantoms[i]
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell = sheet.cell(max_row + 1 + dim, j)
            cell.value = "TOT"
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            print("ID Done!")

        if HEADER.value == "MAE":
            for i in range(dim):
                cell = sheet.cell(max_row + 1 + i, j)
                cell.value = MAE_list[i]
            mean_cell = sheet.cell(max_row + 1 + dim, j)
            mean_cell.value = np.mean(MAE_list)
            std_cell = sheet.cell(max_row + 2 + dim, j)
            std_cell.value = np.std(MAE_list)
            print("MAE Done!")

        if HEADER.value == "SSIM":
            for i in range(dim):
                cell = sheet.cell(max_row + 1 + i, j)
                cell.value = SSIM_list[i]
            mean_cell = sheet.cell(max_row + 1 + dim, j)
            mean_cell.value = np.mean(SSIM_list)
            std_cell = sheet.cell(max_row + 2 + dim, j)
            std_cell.value = np.std(SSIM_list)
            print("SSIM Done!")

        if HEADER.value == "FSIM":
            for i in range(dim):
                cell = sheet.cell(max_row + 1 + i, j)
                cell.value = FSIM_list[i]
            mean_cell = sheet.cell(max_row + 1 + dim, j)
            mean_cell.value = np.mean(FSIM_list)
            std_cell = sheet.cell(max_row + 2 + dim, j)
            std_cell.value = np.std(FSIM_list)
            print("FSIM Done!")

        if HEADER.value == "EPR":
            for i in range(dim):
                cell = sheet.cell(max_row + 1 + i, j)
                cell.value = EPR_list[i]
            mean_cell = sheet.cell(max_row + 1 + dim, j)
            mean_cell.value = np.mean(EPR_list)
            std_cell = sheet.cell(max_row + 2 + dim, j)
            std_cell.value = np.std(EPR_list)
            print("EPR Done!")

        if HEADER.value == "EGR":
            for i in range(dim):
                cell = sheet.cell(max_row + 1 + i, j)
                cell.value = EGR_list[i]
            mean_cell = sheet.cell(max_row + 1 + dim, j)
            mean_cell.value = np.mean(EGR_list)
            std_cell = sheet.cell(max_row + 2 + dim, j)
            std_cell.value = np.std(EGR_list)
            print("EGR Done!")
    
    wb_obj.save(path)

def save_xlsx_unpaired(path, id_phantoms, id_patients, HistCC_list, NCC_list, MEAN_liver_ph_list, NM_liver_ph_list, MEAN_liver_pz_list, NM_liver_pz_list):
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet = wb_obj.active  # Access the active sheet
    max_column = sheet.max_column
    max_row = sheet.max_row
    dim = len(id_phantoms)

    for j in range(1, max_column + 1):
        HEADER = sheet.cell(row=1, column=j)

        if HEADER.value == "ID":
            for i in range(dim):
                cell = sheet.cell(max_row + 1 + i, j)
                cell.value = str(id_phantoms[i]) + '-' + str(id_patients[i]) 
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell = sheet.cell(max_row + 1 + dim, j)
            cell.value = "TOT"
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            print("ID Done!")

        if HEADER.value == "HistCC":
            for i in range(dim):
                cell = sheet.cell(max_row + 1 + i, j)
                cell.value = HistCC_list[i]
            mean_cell = sheet.cell(max_row + 1 + dim, j)
            mean_cell.value = np.mean(HistCC_list)
            std_cell = sheet.cell(max_row + 2 + dim, j)
            std_cell.value = np.std(HistCC_list)
            print("HistCC Done!")

        if HEADER.value == "NCC":
            for i in range(dim):
                cell = sheet.cell(max_row + 1 + i, j)
                cell.value = NCC_list[i]
            mean_cell = sheet.cell(max_row + 1 + dim, j)
            mean_cell.value = np.mean(NCC_list)
            std_cell = sheet.cell(max_row + 2 + dim, j)
            std_cell.value = np.std(NCC_list)
            print("NCC Done!")

        if HEADER.value == "Mean Liver ph":
            for i in range(dim):
                cell = sheet.cell(max_row + 1 + i, j)
                cell.value = MEAN_liver_ph_list[i]
            mean_cell = sheet.cell(max_row + 1 + dim, j)
            mean_cell.value = np.mean(MEAN_liver_ph_list)
            std_cell = sheet.cell(max_row + 2 + dim, j)
            std_cell.value = np.std(MEAN_liver_ph_list)
            print("Mean Liver ph Done!")

        if HEADER.value == "NM Liver ph":
            for i in range(dim):
                cell = sheet.cell(max_row + 1 + i, j)
                cell.value = NM_liver_ph_list[i]
            mean_cell = sheet.cell(max_row + 1 + dim, j)
            mean_cell.value = np.mean(NM_liver_ph_list)
            std_cell = sheet.cell(max_row + 2 + dim, j)
            std_cell.value = np.std(NM_liver_ph_list)
            print("NM Liver ph Done!")

        if HEADER.value == "Mean Liver pz":
            for i in range(dim):
                cell = sheet.cell(max_row + 1 + i, j)
                cell.value = MEAN_liver_pz_list[i]
            mean_cell = sheet.cell(max_row + 1 + dim, j)
            mean_cell.value = np.mean(MEAN_liver_pz_list)
            std_cell = sheet.cell(max_row + 2 + dim, j)
            std_cell.value = np.std(MEAN_liver_pz_list)
            print("Mean Liver pz Done!")
        
        if HEADER.value == "NM Liver pz":
            for i in range(dim):
                cell = sheet.cell(max_row + 1 + i, j)
                cell.value = NM_liver_pz_list[i]
            mean_cell = sheet.cell(max_row + 1 + dim, j)
            mean_cell.value = np.mean(NM_liver_pz_list)
            std_cell = sheet.cell(max_row + 2 + dim, j)
            std_cell.value = np.std(NM_liver_pz_list)
            print("NP Liver pz Done!")
    
    wb_obj.save(path)

def radial_nps(axial_slices, pixel_spacing):
    radial_nps_list = []
    for slice in axial_slices:
        
        I_G = gaussian_filter(slice, sigma=1) # gaussian filtering
        detrended_slice = slice - I_G # detrend
        fft_data = fft2(detrended_slice) # compute 2D FFT
        power_spectrum = np.abs(fft_data) ** 2 # compute power spectrum NPS

        L_G = np.sum(I_G > 0)  # Number of nonzero pixels in the Gaussian-filtered slice
        normalized_ps = power_spectrum / (pixel_spacing[0] * pixel_spacing[1] * L_G) # normalize NPS
        #print(normalized_ps.shape)
        #plt.imshow(normalized_ps)
        #plt.show()
        
        center = tuple(np.array(normalized_ps.shape) // 2) # radial averaging
        y, x = np.indices(normalized_ps.shape)
        r = np.sqrt((x - center[1])**2 + (y - center[0])**2)
        r = r.astype(int)

        tbin = np.bincount(r.ravel(), normalized_ps.ravel())
        nr = np.bincount(r.ravel())
        radial_nps = tbin / nr
        
        radial_nps_list.append(radial_nps[:normalized_ps.shape[0] // 2])
    
    radial_nps_avg = np.mean(radial_nps_list, axis=0) # average over all slices
    return radial_nps_avg

def ncc(axial_slices_synthetic, axial_slices_patient, pixel_spacing):
    
    # Compute NPS for synthetic and patient data
    nps_synthetic = radial_nps(axial_slices_synthetic, pixel_spacing)
    nps_patient = radial_nps(axial_slices_patient, pixel_spacing)
    # Calculate Pearson correlation coefficient (NCC) between patient and synthetic
    ncc, _ = pearsonr(nps_synthetic, nps_patient)
    
    return nps_synthetic, nps_patient, ncc

def metrics_unpaired(volume_true, volume_pred, liver_mask_pred, liver_mask_pz, roi_ph, roi_pz, pixel_spacing):
    
    # HistCC
    counts_true = np.histogram(volume_true.numpy(), bins=100)
    counts_pred = np.histogram(volume_pred.numpy(), bins=100)
    correlation = np.corrcoef(counts_true[0], counts_pred[0])
    HistCC = correlation[1,0]

    # NCC
    nps_ph, nps_pz, NCC = ncc(roi_ph, roi_pz, pixel_spacing)
    
    # NM liver fakeMRI
    mean_liver_ph = np.mean(volume_pred[liver_mask_pred==1])
    std_liver_ph = np.std(volume_pred[liver_mask_pred==1])
    
    # NM liver realMRI
    mean_liver_pz = np.mean(volume_true[liver_mask_pz==1])
    std_liver_pz = np.std(volume_true[liver_mask_pz==1])

    return HistCC, mean_liver_ph, std_liver_ph, mean_liver_pz, std_liver_pz, NCC, nps_ph, nps_pz

def pad_to_shape(slice_img, target_shape=(348, 348)):
    slice_img = np.squeeze(slice_img)
    h, w = slice_img.shape
    target_h, target_w = target_shape

    pad_top = (target_h - h) // 2
    pad_bottom = target_h - h - pad_top
    pad_left = (target_w - w) // 2
    pad_right = target_w - w - pad_left

    padded = np.pad(slice_img, ((pad_top, pad_bottom), (pad_left, pad_right)), mode='constant', constant_values=0)
    padded = tf.expand_dims(padded, axis=2)
    #print(padded.shape)
    return padded


#%%

#____________________________COMPUTE_METRICS_PAIRED_________________________________________________________
# compare realistic phantom with original phantom computing metrics on one-to-one corresponding axial slices

MAE_list = []
SSIM_list = []
FSIM_list = []
EPR_list = []
EGR_list = []

for index in id_ph:
    print('phantom:',index)
    dir_COMBAT = folder_COMBAT + str(index) + '_recon_denoised_clean99_hm.mha'
    COMBAT = load(dir_COMBAT)                                           # + str(desired) +
    dir_fakeMRI = folder_MRI_corrected + str(index) + '_fakeMRI_136_train_corr_prova_320x260_new.mha'
    fakeMRI = load(dir_fakeMRI)
    dir_mask = folder_Mask + str(index) + '_mask.mha'
    vol_mask = load(dir_mask)
    
    SSIM = []
    FSIM = []
    EGR = []
    EPR = []
    MAE = []
    
    assert COMBAT.shape[2]==fakeMRI.shape[2]==vol_mask.shape[2], "No matching dimensions"
    
    for slice in range(0, fakeMRI.shape[2]):
        slice_combat = extract_slice(COMBAT, slice)
        slice_fakeMRI = extract_slice(fakeMRI, slice)
        slice_mask = extract_slice(vol_mask, slice)
    #for slice in range(0, fakeMRI.shape[2]):
        #print(extract_slice(COMBAT, slice))
        # slice_combat = pad_to_shape(extract_slice(COMBAT, slice))
        # slice_fakeMRI = pad_to_shape(extract_slice(fakeMRI, slice))
        # slice_mask = pad_to_shape(extract_slice(vol_mask, slice))
        
        #assert slice_combat.shape == (348, 348, 1)
        #assert slice_fakeMRI.shape == (348, 348, 1)
        #assert slice_mask.shape == (348, 348, 1)

        
        mae = mean_absolute_error(slice_fakeMRI, slice_combat, slice_mask)
        MAE.append(mae)
        
        slice_combat = normalize(slice_combat)
        slice_fakeMRI = normalize(slice_fakeMRI)
        
        # SSIM and FSIM
        ssim_value = ssim(slice_fakeMRI.numpy(), slice_combat.numpy(), 2)
        fsim_value = fsim(slice_fakeMRI.numpy(), slice_combat.numpy(), 2)
        SSIM.append(ssim_value)
        FSIM.append(fsim_value)
        
        # EGR and EPR (otherwise use function edge_ratios)
        img1 = tf.convert_to_tensor(slice_combat, dtype=tf.float32)
        img2 = tf.convert_to_tensor(slice_fakeMRI, dtype=tf.float32)
        sobel_filter = tf.constant([[1, 2, 1], [0, 0, 0], [-1, -2, -1]], dtype=tf.float32)
        img1_edges = tf.nn.conv2d(np.expand_dims(img1, 0), tf.expand_dims(tf.expand_dims(sobel_filter, -1), -1), strides=[1, 1, 1, 1], padding='SAME')
        img2_edges = tf.nn.conv2d(np.expand_dims(img2, 0), tf.expand_dims(tf.expand_dims(sobel_filter, -1), -1), strides=[1, 1, 1, 1], padding='SAME')
        edge_gt = np.abs(tf.squeeze(img1_edges))
        edge_pred = np.abs(tf.squeeze(img2_edges))
        threshold = 0.8
        _, binary_img_gt = cv2.threshold(edge_gt, threshold, 1, cv2.THRESH_BINARY)
        _, binary_img_pred = cv2.threshold(edge_pred, threshold, 1, cv2.THRESH_BINARY)
        inter = cv2.bitwise_and(binary_img_gt, binary_img_pred) # edge map in common between edges1 and edges2
        inter = np.abs(inter)
        _, binary_img_inter = cv2.threshold(inter, threshold * 0.625, 1, cv2.THRESH_BINARY)  # 0.625 è un valore approssimato per mantenere l'intersezione valida
        tot_gt = np.sum(binary_img_gt)
        tot_pred = np.sum(binary_img_pred)
        tot_inter = np.sum(binary_img_inter)
        epr = tot_inter / tot_gt
        egr = tot_pred / tot_gt
        EPR.append(epr)
        EGR.append(egr)
    
    # # histogram of each metric for each patient
    # metrics = [SSIM, FSIM, EPR, EGR, MAE]
    # labels = ['SSIM', 'FSIM', 'EPR', 'EGR', 'MAE']
    # colors = ['lightblue', 'lightgreen', 'red', 'violet', 'orange']
    # xlims = [(0, 1), (0.5, 1), (0.4, 1), (0.5, 1.5), (10, 30)]  
    # num_metrics = len(metrics)
    # font_label = {'family': 'serif', 'color': 'black', 'weight': 'normal', 'size': 14}
    # font_title = {'family': 'serif', 'color': 'black', 'weight': 'normal', 'size': 16}
    
    # fig, axes = plt.subplots(1, num_metrics, figsize=(7* num_metrics, 10))
    # for i, ax in enumerate(axes):
    #     ax.hist(metrics[i], bins=20, color=colors[i], edgecolor='black', alpha=0.7)
    #     ax.set_xlim(xlims[i])
    #     mean_value = np.mean(metrics[i])
    #     ax.axvline(mean_value, color='black', linestyle='dashed', linewidth=2, label=f'Mean = {mean_value:.3f}')

    #     ax.set_xlabel('Metric Value', fontdict=font_label)
    #     ax.set_ylabel('Frequency', fontdict=font_label)
    #     ax.set_title(labels[i] + ' distribution', fontdict=font_title)
    #     if i!=4:
    #         ax.legend(loc='upper left')
    #     else:
    #         ax.legend(loc='upper right')
    
    # plt.suptitle('Distribution of Metrics for phantom ' + str(index), fontsize=18, fontweight='bold', fontfamily='serif')
    # plt.savefig('//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/MRI/' + desired + '/Metrics_distributions_' + desired + str(index) + '.png')
    # plt.show()

    MAE_list.append(np.mean(MAE))
    SSIM_list.append(np.mean(SSIM))
    FSIM_list.append(np.mean(FSIM))
    EPR_list.append(np.mean(EPR))
    EGR_list.append(np.mean(EGR))
    
# save metrics in xslx file
#path = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/MRI/' + desired + '/fakeMRI_paired_' + desired + '_metrics_320x260.xlsx'
path = '//NAS-CARTCAS/camagnif/tesi/Data/mri/results_revisori/5th/test/inference/test_volumes/fakeMRI_paired_test_metrics_136_320x260.xlsx'
save_xlsx_paired(path, id_ph, MAE_list, SSIM_list, FSIM_list, EPR_list, EGR_list)
       
print('SSIM medio vale ', np.mean(SSIM_list), ' con std ', np.std(SSIM_list))
print('FSIM medio vale ', np.mean(FSIM_list), ' con std ', np.std(FSIM_list))
print('EPR medio vale ', np.mean(EPR_list), ' con std ', np.std(EPR_list))
print('EGR medio vale ', np.mean(EGR_list), ' con std ', np.std(EGR_list))
print('MAE medio vale ', np.mean(MAE_list), ' con std ', np.std(MAE_list))

#%%

#____________________________COMPUTE_METRICS_UNPAIRED__________________________
# compare realistic phantom with real patients computing un-paired metrics

id_ph = [50, 71, 86, 99, 143]
id_pz = [9, 14, 21, 16, 30]
list_folders = os.listdir(folder_unpaired)

HistCC = []
MEAN_liver_ph = []
NM_liver_ph = []
MEAN_liver_pz = []
NM_liver_pz = []
NCC = []
pixel_spacing = [1.0625, 1.0625]

for num, id in enumerate(id_ph):
    print('Phantom', id, 'and Unpaired patient', id_pz[num])
    for folder in list_folders:
        if folder.__contains__(str(id)):
            folder_data = folder_unpaired + folder + "/"
            ph = folder_data + str(id) + '_MRIfake.mha'
            liver_ph = folder_data + str(id) + '_liver.mha'
            roi_ph = folder_data + str(id) + '_roi.mha'
            pz = folder_data + str(id_pz[num]) + '_MRI.mha'
            liver_pz = folder_data + str(id_pz[num]) + '_liver.mha'
            roi_pz = folder_data + str(id_pz[num]) + '_roi.mha'
            
            MRI_fake = load(ph)
            liver_mask_ph = load(liver_ph)
            MRI = load(pz)
            liver_mask_pz = load(liver_pz)
            roi_ph = (load(roi_ph)).numpy()
            roi_pz = (load(roi_pz)).numpy()
    
            histcc, mean_ph, std_ph, mean_pz, std_pz, ncc_coeff, nps_ph, nps_pz = metrics_unpaired(MRI, MRI_fake, liver_mask_ph, liver_mask_pz, roi_ph, roi_pz, pixel_spacing)
            
            print('Histogram correlation between fakeMRI and real MRI is: ', histcc)
            print('Mean value inside realistic phantom liver is: ' , mean_ph, 'with std', std_ph)
            print('Mean value inside real patient liver is: ', mean_pz, 'with std', std_pz)
            print("NPS Correlation Coefficient (NCC):", ncc_coeff)
            
            HistCC.append(histcc)
            MEAN_liver_ph.append(mean_ph)
            NM_liver_ph.append(std_ph)
            MEAN_liver_pz.append(mean_pz)
            NM_liver_pz.append(std_pz)
            NCC.append(ncc_coeff)
            
            # Plot Radial NPS
            # frequencies = np.arange(len(nps_ph)) * (1 / (pixel_spacing[0] * len(nps_ph)))
            
            # plt.figure(figsize=(10, 6))
            # plt.plot(frequencies, nps_ph, label='Radial NPS - Phantom')
            # plt.plot(frequencies, nps_pz, label='Radial NPS - Patient')
            # plt.xlabel('Frequency (cycles/mm)')
            # plt.ylabel('Radial NPS')
            # plt.title('Comparison of Radial NPS between Phantom ' + str(id) +' and Patient ' + str(id_pz[num]))
            # plt.legend()
            # #plt.grid(True)
            # plt.savefig(folder_data + '/RadialNPS_ph' + str(id) + '_pz' + str(id_pz[num]))
                
# save metrics in xslx file
path = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/MRI/test/fakeMRI_unpaired_' + desired + '_metrics.xlsx'
save_xlsx_unpaired(path, id_ph, id_pz, HistCC, NCC, MEAN_liver_ph, NM_liver_ph, MEAN_liver_pz, NM_liver_pz)
       
print('HistCC medio vale ', np.mean(HistCC), ' con std ', np.std(HistCC))
print('NCC medio vale ', np.mean(NCC), ' con std ', np.std(NCC))
print('MEAN Liver phantom medio vale ', np.mean(MEAN_liver_ph), ' con std ', np.std(MEAN_liver_ph))
print('NM Liver phantom medio vale ', np.mean(NM_liver_ph), ' con std ', np.std(NM_liver_ph))
print('MEAN Liver patient medio vale ', np.mean(MEAN_liver_pz), ' con std ', np.std(MEAN_liver_pz))
print('NM Liver patient medio vale ', np.mean(NM_liver_pz), ' con std ', np.std(NM_liver_pz))

#%% 
#____________________________________HISTOGRAMS________________________________
# average histogram of 5 patients MRI vs average histogram of 5 realistic fake MRI phantoms

MRI_fake = glob.glob(folder_all_fake + '*.mha')
MRI_real = glob.glob(folder_all_pz + '*.mha')
ph = glob.glob(folder_phantom + '*.mha')
assert len(MRI_fake) == len(MRI_real) == len(ph), 'No matching dimensions'

bins = 20
hist_sum_fake = np.zeros(bins)
hist_sum_real = np.zeros(bins)
#hist_sum = np.zeros(bins)
hist_sum = np.zeros(80)

# FAKE
for i in range(len(MRI_fake)): 
    MRI = load(MRI_fake[i])
    #MRI = MRI.flatten()
    hist, bin_edges_f = np.histogram(MRI, bins=bins, range=(0, 100))
    hist_normalized = hist / np.sum(hist)
    hist_sum_fake += hist_normalized
# Mean histogram
mean_hist_fake = hist_sum_fake / len(MRI_fake)

# REAL PZ
for i in range(len(MRI_real)): 
    MRI = load(MRI_real[i])
    #MRI = MRI.flatten()
    hist, bin_edges_r = np.histogram(MRI, bins=bins, range=(0, 100))
    hist_normalized = hist / np.sum(hist)
    hist_sum_real += hist_normalized
# Mean histogram
mean_hist_real = hist_sum_real / len(MRI_real)

# PHANTOMS
min_old_ph = 0
max_old_ph = 140
new_min_ph = 0
new_max_ph = 90

for i in range(len(ph)): 
    phantom = load(ph[i])
    #phantom = phantom.flatten()
    phantom_normalized = (phantom - min_old_ph) * (new_max_ph - new_min_ph) / (max_old_ph - min_old_ph) + new_min_ph
    hist, bin_edges_p = np.histogram(phantom_normalized, bins=80, range=(0, 96))
    hist_normalized = hist / np.sum(hist)                #80
    hist_sum += hist_normalized
# Mean histogram
mean_hist = hist_sum / len(ph)

# Plot 
plt.figure(figsize=(15, 10))
plt.rcParams['font.family'] = 'serif'
plt.plot(bin_edges_f[:-1], mean_hist_fake, color='coral', linestyle='-', linewidth=2, label='Realistic MRI')
plt.fill_between(bin_edges_f[:-1], mean_hist_fake, color='coral', alpha=0.3, edgecolor='coral')
plt.plot(bin_edges_r[:-1], mean_hist_real, color='darkviolet', linestyle='-', linewidth=2, label='Real MRI')
plt.fill_between(bin_edges_r[:-1], mean_hist_real, color='darkviolet', alpha=0.3, edgecolor='darkviolet')
plt.plot(bin_edges_p[:-1], mean_hist, color='mediumseagreen', linestyle='-', linewidth=2, label='CoMBAT')
#plt.fill_between(bin_edges_p[:-1], mean_hist, color='mediumseagreen', alpha=0.3, edgecolor='mediumseagreen')
plt.ylabel('Relative frequency', fontsize=20)
plt.xlabel('Voxel Intensity [a.u.]', fontsize=20)
plt.xticks(fontsize=20)
plt.yticks(fontsize=20)
plt.legend(fontsize=20)
plt.title('Realistic MRI vs Real MRI vs CoMBAT', fontsize=24)
plt.tight_layout()
plt.savefig(folder + 'fakeMRI_vs_realMRI_vs_COMBAT_histogram_comparison_80binCOMBATinsteadof20.png')
plt.show()

#%%

#________________________PREDICTION_ON_PATCHES_________________________________
# since the net has been trained on patches, metrics should be computed on patches:
# therefore:
#   1. the test image (348x348) is loaded
#   2. a random patch of dimensions 220x220 is extracted (according to training)
#   3. resize each patch to 256x256 for model prediction and resize back to 220x220
#   4. metrics computation on corresponding patches between original_ph patch and realistic_ph patch

#_____________________________FUNCTIONS________________________________________

RESIZE_DIM = 256

def expand_dim(input_img, axis):
    output = tf.expand_dims(input_img, axis) 
    return output

def resize(input_image):                        #NOT CHANGE: 256!
    input_image = tf.image.resize(input_image, [RESIZE_DIM, RESIZE_DIM], method=tf.image.ResizeMethod.NEAREST_NEIGHBOR, preserve_aspect_ratio=False)
    return input_image

def resize_new(input_image):                    #CHANGE
    input_image = tf.image.resize(input_image, [220, 220], method=tf.image.ResizeMethod.NEAREST_NEIGHBOR, preserve_aspect_ratio=False)
    return input_image

def normalize(input_image):
    input_image = 2*((input_image-np.min(input_image)) / (np.max(input_image)-np.min(input_image))) - 1
    return input_image

# from [-1;+1] back to original values range: [0-100] for MRI VIBE
def merge_HU(predTensor):
    ct = predTensor[:, :]
    ctnew = 0.5*(100-(0))*(ct+1) + (0)  
    return ctnew

#__________________________________CYCLEGAN____________________________________

input_img_size = (256,256,1)

class ReflectionPadding2D(layers.Layer):

    def __init__(self, padding=(1, 1), **kwargs):
        self.padding = tuple(padding)
        super(ReflectionPadding2D, self).__init__(**kwargs)

    def call(self, input_tensor, mask=None):
        padding_width, padding_height = self.padding
        padding_tensor = [
            [0, 0],
            [padding_height, padding_height],
            [padding_width, padding_width],
            [0, 0],
        ]
        return tf.pad(input_tensor, padding_tensor, mode="REFLECT")

# Weights initializer for the layers
kernel_init = keras.initializers.RandomNormal(mean=0.0, stddev=0.02)
# Gamma initializer for instance normalization
gamma_init = keras.initializers.RandomNormal(mean=0.0, stddev=0.02)


def residual_block(x, activation, kernel_initializer=kernel_init, kernel_size=(3, 3), strides=(1, 1), padding="valid", gamma_initializer=gamma_init, use_bias=False):
    dim = x.shape[-1]
    input_tensor = x

    x = ReflectionPadding2D()(input_tensor)
    x = layers.Conv2D(dim, kernel_size, strides=strides, kernel_initializer=kernel_initializer, padding=padding, use_bias=use_bias)(x)
    x = tfa.layers.InstanceNormalization(gamma_initializer=gamma_initializer)(x)
    x = activation(x)

    x = ReflectionPadding2D()(x)
    x = layers.Conv2D(dim, kernel_size, strides=strides, kernel_initializer=kernel_initializer, padding=padding, use_bias=use_bias)(x)
    x = tfa.layers.InstanceNormalization(gamma_initializer=gamma_initializer)(x)
    x = layers.add([input_tensor, x])
    return x


def downsample(x, filters, activation, kernel_initializer=kernel_init, kernel_size=(3, 3), strides=(2, 2), padding="same", gamma_initializer=gamma_init, use_bias=False): 
    x = layers.Conv2D(filters, kernel_size, strides=strides, kernel_initializer=kernel_initializer, padding=padding, use_bias=use_bias)(x)
    x = tfa.layers.InstanceNormalization(gamma_initializer=gamma_initializer)(x)
    if activation:
        x = activation(x)
    return x


def upsample(x, filters, activation, kernel_size=(3, 3), strides=(1, 1), padding="valid", kernel_initializer=kernel_init, gamma_initializer=gamma_init, use_bias=False):
    #dim = x.shape[-1]
    x = layers.UpSampling2D(size=(2, 2), interpolation="bilinear")(x)
    x = ReflectionPadding2D()(x)
    x = layers.Conv2D(filters, kernel_size, strides=strides, kernel_initializer=kernel_initializer, padding=padding, use_bias=use_bias)(x)
    x = tfa.layers.InstanceNormalization(gamma_initializer=gamma_initializer)(x)
    x = activation(x)
    return x


def get_resnet_generator(filters=64, num_downsampling_blocks=2, num_residual_blocks=9, num_upsample_blocks=2, gamma_initializer=gamma_init):
    img_input = layers.Input(shape=input_img_size)
    #mask = layers.Input(shape=input_img_size)
    x = ReflectionPadding2D(padding=(3, 3))(img_input)
    x = layers.Conv2D(filters, (7, 7), kernel_initializer=kernel_init, use_bias=False)(x)
    x = tfa.layers.InstanceNormalization(gamma_initializer=gamma_initializer)(x)
    x = layers.Activation("relu")(x)

    # Downsampling
    for _ in range(num_downsampling_blocks):
        filters *= 2
        x = downsample(x, filters=filters, activation=layers.Activation("relu"))

    # Residual blocks
    for _ in range(num_residual_blocks):
        x = residual_block(x, activation=layers.Activation("relu"))

    # Upsampling
    for _ in range(num_upsample_blocks):
        filters //= 2
        x = upsample(x, filters=filters, activation=layers.Activation("relu"))

    # Final block
    x = ReflectionPadding2D(padding=(3, 3))(x)
    x = layers.Conv2D(1, (7, 7), padding="valid")(x)
    x = layers.Activation("tanh")(x)
    #x = Multiply()([x, mask])

    model = keras.models.Model(img_input, x)
    return model

#_________________________________INITIALIZATION_______________________________

input_img_size = (256,256,1)

# Get the generators
gen_G = get_resnet_generator()
gen_F = get_resnet_generator()

# Plot G e D architectures
tf.keras.utils.plot_model(gen_G, show_shapes=True, dpi=64)

# Define the the optimizers for the generator and the discriminator
LR_gen = 2e-4  # Learning rate for the generator
LR_disc = 2e-4 # Learning rate for the discriminator
generator_g_optimizer = tf.keras.optimizers.Adam(learning_rate = LR_gen, beta_1=0.5)
generator_f_optimizer = tf.keras.optimizers.Adam(learning_rate = LR_gen, beta_1=0.5)

# Load the best generator
best_epoch = 150 # best epoch
gen_G.load_weights(folder_weights +'/generator_g_' + str(int(best_epoch)) + '.h5')  
gen_F.load_weights(folder_weights +'/generator_f_' + str(int(best_epoch)) + '.h5')  

#%%

#_____________________________METRICS_ON_PATCHES_______________________________

MAE_list = []
SSIM_list = []
FSIM_list = []
EPR_list = []
EGR_list = []

for index in id_ph:
    print('phantom:',index)
    dir_COMBAT = folder_COMBAT + str(index) + '_recon_denoised_clean99_hm.mha'
    COMBAT = load(dir_COMBAT)
    dir_mask = folder_Mask + str(index) + '_mask.mha'
    vol_mask = load(dir_mask)
    
    SSIM = []
    FSIM = []
    EGR = []
    EPR = []
    MAE = []
    size = 220
    stride = 32
    model = gen_G
    
    assert COMBAT.shape[2]==vol_mask.shape[2], "No matching dimensions"
    
    for slice in range(0, COMBAT.shape[2]):
        
        slice_combat = expand_dim(extract_slice(COMBAT, slice), 0) # [1,348, 348, 1]
        slice_mask = expand_dim(extract_slice(vol_mask, slice), 0)
        slice_combat = normalize(slice_combat)
        
        patches_combat = tf.image.extract_patches(images = slice_combat, sizes=[1, size, size, 1],
                                   strides=[1, stride, stride, 1], rates=[1, 1, 1, 1], padding='VALID')
        n_patches = patches_combat.shape[1]*patches_combat.shape[2]
        patch_combat_all = tf.reshape(patches_combat, shape=(n_patches, size, size, 1))
        
        patches_mask = tf.image.extract_patches(images = slice_mask, sizes=[1, size, size, 1],
                                   strides=[1, stride, stride, 1], rates=[1, 1, 1, 1], padding='VALID')
        n_patches = patches_mask.shape[1]*patches_mask.shape[2]
        patch_mask_all = tf.reshape(patches_mask, shape=(n_patches, size, size, 1)) # [25, 220, 220, 1]
        
        index = randint(0, n_patches-1)
        patch_combat_sel = patch_combat_all[index, :,:,:]
        patch_mask_sel = patch_mask_all[index, :,:,:]
        patch_temp = expand_dim(patch_combat_sel, 0) #patch: [220, 220, 1] patch_temp: [1, 220, 220, 1]
        tensor = expand_dim(tf.squeeze(resize_new(model(resize(patch_temp)))),0)
        
        plt.imshow(tf.squeeze(tensor))
        plt.show()
        
        pred_patch = expand_dim(merge_HU(tf.squeeze(tensor)), 2)
        patch_combat_sel = patch_combat_sel.numpy()
        patch_combat_sel[patch_mask_sel==0]=0
        gt_patch = patch_combat_sel
        mask_patch = patch_mask_sel
        pred_patch = pred_patch.numpy()
        pred_patch[mask_patch==0]=0
        
        mae = mean_absolute_error(pred_patch, gt_patch, mask_patch)
        MAE.append(mae)
        
        pred_patch = normalize(pred_patch)
        gt_patch = normalize(gt_patch)
    
        # SSIM and FSIM
        ssim_value = ssim(pred_patch, gt_patch, 2)
        fsim_value = fsim(pred_patch, gt_patch, 2)
        SSIM.append(ssim_value)
        FSIM.append(fsim_value)
        
        # EGR and EPR (otherwise use function edge_ratios)
        img1 = tf.convert_to_tensor(gt_patch, dtype=tf.float32)
        img2 = tf.convert_to_tensor(pred_patch, dtype=tf.float32)
        sobel_filter = tf.constant([[1, 2, 1], [0, 0, 0], [-1, -2, -1]], dtype=tf.float32)
        img1_edges = tf.nn.conv2d(np.expand_dims(img1, 0), tf.expand_dims(tf.expand_dims(sobel_filter, -1), -1), strides=[1, 1, 1, 1], padding='SAME')
        img2_edges = tf.nn.conv2d(np.expand_dims(img2, 0), tf.expand_dims(tf.expand_dims(sobel_filter, -1), -1), strides=[1, 1, 1, 1], padding='SAME')
        edge_gt = np.abs(tf.squeeze(img1_edges))
        edge_pred = np.abs(tf.squeeze(img2_edges))
        threshold = 0.8
        _, binary_img_gt = cv2.threshold(edge_gt, threshold, 1, cv2.THRESH_BINARY)
        _, binary_img_pred = cv2.threshold(edge_pred, threshold, 1, cv2.THRESH_BINARY)
        inter = cv2.bitwise_and(binary_img_gt, binary_img_pred) # edge map in common between edges1 and edges2
        inter = np.abs(inter)
        _, binary_img_inter = cv2.threshold(inter, threshold * 0.625, 1, cv2.THRESH_BINARY)  # 0.625 è un valore approssimato per mantenere l'intersezione valida
        tot_gt = np.sum(binary_img_gt)
        tot_pred = np.sum(binary_img_pred)
        tot_inter = np.sum(binary_img_inter)
        epr = tot_inter / tot_gt
        egr = tot_pred / tot_gt
        EPR.append(epr)
        EGR.append(egr)

    MAE_list.append(np.mean(MAE))
    SSIM_list.append(np.mean(SSIM))
    FSIM_list.append(np.mean(FSIM))
    EPR_list.append(np.mean(EPR))
    EGR_list.append(np.mean(EGR))
    
# save metrics in xslx file
path = '//NAS-CARTCAS/camagnif/tesi/FINAL_RESULTS/MRI/' + desired + '/fakeMRI_paired_patches_' + desired + '_metrics_348x348.xlsx'
save_xlsx_paired(path, id_ph, MAE_list, SSIM_list, FSIM_list, EPR_list, EGR_list)
       
print('SSIM medio vale ', np.mean(SSIM_list), ' con std ', np.std(SSIM_list))
print('FSIM medio vale ', np.mean(FSIM_list), ' con std ', np.std(FSIM_list))
print('EPR medio vale ', np.mean(EPR_list), ' con std ', np.std(EPR_list))
print('EGR medio vale ', np.mean(EGR_list), ' con std ', np.std(EGR_list))
print('MAE medio vale ', np.mean(MAE_list), ' con std ', np.std(MAE_list))



